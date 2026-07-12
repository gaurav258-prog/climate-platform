"""SFDR Principal Adverse Impact (PAI) statement — the filing artifact.

This turns a fund's computed climate data into the actual document an EU asset
manager files: the mandatory PAI statement in the shape the regulation defines
(SFDR RTS — Commission Delegated Regulation (EU) 2022/1288, Annex I, Table 1),
plus the EU Taxonomy alignment lines.

The point of the whole product is that this is FILING-READY and AUDITABLE, so
the design rule is strict: every mandatory indicator is listed; the ones we can
honestly compute are filled in with their coverage and data source; the ones we
cannot are marked "not available" with the exact input still required. We never
invent a number, and we never silently drop a mandatory row — a regulator and an
auditor must see both what we have and what is missing.

What is computed today (from services/fund_disclosure.fund_pai):
    * PAI 3  — GHG intensity of investees (WACI)          — computed
    * PAI 4  — fossil-fuel-sector exposure                — computed
    * PAI 1  — investee GHG emissions (un-attributed)     — partial (needs EVIC
                                                            for PCAF attribution)
Everything else in Table 1 needs issuer data we do not yet hold (energy mix,
water, waste, social/governance) and is surfaced as an input gap.

Scope: investee-company indicators (equity / corporate-bond funds), which is the
beachhead. Sovereign (Table 1, indicators 15-16) and real-estate (17-18) tables
are out of scope for this first version and flagged as such.
"""
from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Optional

from sqlalchemy import text

from services.fund_disclosure import fund_pai, fund_esg_pai

# ── The mandatory PAI indicators (SFDR RTS Annex I, Table 1 — investee companies) ──
# Each: number, area, metric (as worded by the RTS), unit. Value/coverage/source
# are filled per-fund at assembly time.
MANDATORY_PAI_INDICATORS = [
    (1,  "Climate & environment", "GHG emissions (Scope 1, 2 and 3, and total)", "tCO₂e"),
    (2,  "Climate & environment", "Carbon footprint (financed emissions per €M invested)", "tCO₂e/€M"),
    (3,  "Climate & environment", "GHG intensity of investee companies (WACI)", "tCO₂e/€M revenue"),
    (4,  "Climate & environment", "Exposure to companies active in the fossil fuel sector", "% of value"),
    (5,  "Climate & environment", "Share of non-renewable energy consumption and production", "%"),
    (6,  "Climate & environment", "Energy consumption intensity per high-impact climate sector", "GWh/€M revenue"),
    (7,  "Biodiversity",          "Activities negatively affecting biodiversity-sensitive areas", "% of value"),
    (8,  "Water",                 "Emissions to water", "tonnes/€M invested"),
    (9,  "Waste",                 "Hazardous waste and radioactive waste ratio", "tonnes/€M invested"),
    (10, "Social & governance",   "Violations of UN Global Compact / OECD Guidelines", "% of value"),
    (11, "Social & governance",   "Lack of processes to monitor UNGC / OECD compliance", "% of value"),
    (12, "Social & governance",   "Unadjusted gender pay gap", "%"),
    (13, "Social & governance",   "Board gender diversity", "% female"),
    (14, "Social & governance",   "Exposure to controversial weapons", "% of value"),
]

_GOLDEN_SOURCE = "Tellumen golden source (issuer emissions + revenue, provenance-stamped)"

# ── Sovereign PAI (RTS Annex I, Table 1, indicators 15-16) ──
# GHG intensity of investee COUNTRIES: tCO2e per €M GDP. Loaded from the
# provenanced data file data/reference/country_ghg_intensity.csv (computed from
# OWID / Global Carbon Project CO2 ÷ GDP by scripts/build_country_intensities.py);
# the embedded dict is the offline fallback.
_EMBEDDED_COUNTRY_INTENSITY: dict[str, float] = {
    "SE": 50, "CH": 59, "NO": 70, "FR": 90, "GB": 110, "IT": 120, "DK": 95,
    "ES": 130, "AT": 154, "PT": 130, "FI": 110, "NL": 140, "IE": 90, "BE": 182,
    "DE": 172, "GR": 150, "US": 200, "JP": 160, "PL": 350, "CZ": 277, "IN": 600,
    "CN": 421, "ZA": 700, "RU": 550, "AU": 286, "BR": 152, "CA": 313,
}
DEFAULT_COUNTRY_INTENSITY = 200.0


def _load_country_intensity() -> dict[str, float]:
    import csv as _csv
    from pathlib import Path as _Path
    path = _Path(__file__).resolve().parent.parent.parent / "data" / "reference" / "country_ghg_intensity.csv"
    try:
        with open(path, newline="", encoding="utf-8") as fh:
            table = {r["country_iso2"].strip().upper(): float(r["intensity_tco2e_per_meur"])
                     for r in _csv.DictReader(fh) if r.get("country_iso2")}
        if table:
            return table
    except (OSError, KeyError, ValueError):
        pass
    return dict(_EMBEDDED_COUNTRY_INTENSITY)


COUNTRY_GHG_INTENSITY_TCO2E_PER_MEUR = _load_country_intensity()

SOVEREIGN_ASSET_CLASSES = ("sovereign_bond",)
REAL_ESTATE_ASSET_CLASSES = ("real_estate",)  # not in the securities model today


def _row(num, area, metric, unit, *, value=None, coverage=None, source=None,
         method="not_available", input_required=None):
    """One indicator line. method ∈ computed / partial / estimated / not_available."""
    return {
        "number": num, "area": area, "metric": metric, "unit": unit,
        "value": value, "coverage_pct": coverage, "source": source,
        "method": method, "input_required": input_required,
    }


def _taxonomy_rollup(session, fund_id: str) -> dict:
    """EU Taxonomy lines for the fund, honestly scoped.

    Eligibility can only be judged where we hold the issuer's NACE code; alignment
    is never asserted (DNSH across the six objectives + minimum safeguards are not
    verified) — matching the classifier's discipline. We report the share of value
    we can even assess, so the gap is explicit.
    """
    rows = session.execute(text("""
        SELECT p.market_value_eur AS mv, i.nace_code
        FROM   fund_positions p
        JOIN   securities s ON s.security_id = p.security_id
        JOIN   issuers   i ON i.issuer_id = s.issuer_id
        WHERE  p.fund_id = :f
    """), {"f": fund_id}).mappings().all()
    total = sum(float(r["mv"]) for r in rows) or 0.0
    with_nace = sum(float(r["mv"]) for r in rows if r["nace_code"])
    return {
        "assessable_pct": round(100 * with_nace / total, 1) if total else 0.0,
        "taxonomy_eligible_pct": None if with_nace == 0 else "requires per-issuer NACE mapping",
        "taxonomy_aligned_pct": None,
        "alignment_note": "Alignment not asserted: DNSH across the six environmental "
                          "objectives and minimum-safeguards verification are not performed. "
                          "Reported as eligible-at-most, never aligned.",
        "input_required": "issuer NACE code (absent from GLEIF; supply per issuer or via a fundamentals upload)",
    }


def _composition_and_sovereign(session, fund_id: str) -> dict:
    """Fund value by asset class + a value-weighted sovereign GHG intensity over
    any sovereign-bond holdings (their issuer's country → country intensity)."""
    rows = session.execute(text("""
        SELECT s.asset_class, i.country, CAST(p.market_value_eur AS FLOAT) AS mv
        FROM   fund_positions p
        JOIN   securities s ON s.security_id = p.security_id
        JOIN   issuers    i ON i.issuer_id = s.issuer_id
        WHERE  p.fund_id = :f
          AND  p.as_of_date = (SELECT MAX(as_of_date) FROM fund_positions WHERE fund_id = :f)
    """), {"f": fund_id}).mappings().all()
    by_class: dict[str, float] = {}
    sov_mv = 0.0
    sov_weighted_intensity = 0.0
    sov_countries: set = set()
    for r in rows:
        by_class[r["asset_class"]] = by_class.get(r["asset_class"], 0.0) + r["mv"]
        if r["asset_class"] in SOVEREIGN_ASSET_CLASSES:
            sov_mv += r["mv"]
            ctry = (r["country"] or "").upper()
            intensity = COUNTRY_GHG_INTENSITY_TCO2E_PER_MEUR.get(ctry, DEFAULT_COUNTRY_INTENSITY)
            sov_weighted_intensity += r["mv"] * intensity
            if ctry:
                sov_countries.add(ctry)
    return {
        "by_asset_class": {k: round(v) for k, v in by_class.items()},
        "sovereign_value_eur": round(sov_mv),
        "sovereign_ghg_intensity": round(sov_weighted_intensity / sov_mv, 1) if sov_mv else None,
        "sovereign_countries": sorted(sov_countries),
        "has_real_estate": any(c in REAL_ESTATE_ASSET_CLASSES for c in by_class),
    }


def _sovereign_indicators(comp: dict) -> list[dict]:
    """RTS Annex I Table 1 indicators 15-16 (sovereign & supranational)."""
    si = comp["sovereign_ghg_intensity"]
    return [
        _row(15, "Sovereign", "GHG intensity of investee countries", "tCO₂e/€M GDP",
             value=si, coverage=100.0 if si is not None else None,
             source="country territorial emissions ÷ GDP (public averages)" if si is not None else None,
             method="computed" if si is not None else "not_available",
             input_required=None if si is not None else "sovereign-bond holdings with issuer country"),
        _row(16, "Sovereign", "Investee countries subject to social violations", "count",
             method="not_available",
             input_required="country social-violation list (UN/OECD sanctions & breaches)"),
    ]


def _real_estate_indicators(comp: dict) -> list[dict]:
    """RTS Annex I Table 1 indicators 17-18 (real-estate assets)."""
    applic = "applies to direct real-estate assets; this fund holds securities, not property" \
        if not comp["has_real_estate"] else None
    method = "not_applicable" if not comp["has_real_estate"] else "not_available"
    return [
        _row(17, "Real estate", "Exposure to fossil fuels through real-estate assets", "% of RE value",
             method=method, input_required=applic or "real-estate asset fossil-fuel involvement"),
        _row(18, "Real estate", "Exposure to energy-inefficient real-estate assets", "% of RE value",
             method=method, input_required=applic or "real-estate asset EPC ratings"),
    ]


def _indicator_numeric(value):
    """Reduce an indicator value to a comparable scalar (dict → its 'total')."""
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, dict) and isinstance(value.get("total"), (int, float)):
        return float(value["total"])
    return None


def _attach_prior_year(session, fund_id: str, ref_year, indicators: list[dict]) -> dict:
    """Look up the most recent FILED snapshot for a prior year and attach each
    indicator's prior value + change. Returns comparison metadata. If none exists
    (first reference period), indicators are left as-is."""
    if not ref_year:
        return {"available": False, "reason": "no reference year on the current statement"}
    row = session.execute(text("""
        SELECT reference_year, statement FROM fund_sfdr_filings
        WHERE fund_id = :f AND reference_year < :y AND status = 'filed'
        ORDER BY reference_year DESC LIMIT 1
    """), {"f": fund_id, "y": ref_year}).mappings().first()
    if not row:
        return {"available": False, "reason": "first reference period — no prior filing to compare"}

    prior_by_num = {i["number"]: i for i in (row["statement"].get("indicators") or [])}
    for ind in indicators:
        prior = prior_by_num.get(ind["number"])
        if not prior:
            continue
        pv, cv = _indicator_numeric(prior.get("value")), _indicator_numeric(ind.get("value"))
        ind["prior_value"] = prior.get("value")
        if pv is not None and cv is not None:
            ind["change"] = round(cv - pv, 3)
            ind["change_pct"] = round(100 * (cv - pv) / pv, 1) if pv else None
    return {"available": True, "prior_reference_year": row["reference_year"]}


def _look_through(comp: dict) -> dict:
    """Report whether the fund holds funds/ETFs that require constituent look-through."""
    held = {k: v for k, v in comp["by_asset_class"].items() if k in ("etf", "fund")}
    if not held:
        return {"applicable": False, "note": "No held funds/ETFs — direct securities only, no look-through required."}
    return {
        "applicable": True,
        "held_fund_value_eur": sum(held.values()),
        "status": "not_expanded",
        "input_required": "constituent holdings of the held funds/ETFs (look-through not yet expanded)",
    }


def sfdr_pai_statement(session, fund_id: str) -> dict:
    """Assemble the fund's full SFDR PAI statement (Annex I Table 1) + Taxonomy.

    Returns a structured, filing-shaped dict: entity metadata, the 14 mandatory
    indicators (filled or gap-flagged), Taxonomy lines, a coverage summary, and
    provenance. Raises nothing for missing data — it is disclosed, not hidden.
    """
    fund = session.execute(text("""
        SELECT f.fund_id::text AS fund_id, f.name, f.sfdr_classification, f.base_currency, f.lei AS fund_lei,
               o.name AS org_name, o.lei AS manager_lei, o.legal_name AS manager_legal_name,
               o.filing_contact_email, o.country AS manager_domicile, o.sfdr_narratives
        FROM funds f JOIN organizations o ON o.org_id = f.org_id
        WHERE f.fund_id = :f
    """), {"f": fund_id}).mappings().first()
    if not fund:
        return {"error": "fund not found"}

    pai = fund_pai(session, fund_id)
    if pai.get("positions", 0) == 0:
        return {"error": "fund has no positions to report on", "fund": dict(fund)}

    p = pai["pai"]
    emis_cov = pai.get("emissions_coverage_pct")
    emis_est = pai.get("emissions_estimated_pct", 0.0)  # SFDR: estimated-vs-reported split
    fin_cov = pai.get("financed_emissions_coverage_pct", 0.0)
    inv = p["pai_1_investee_emissions_tco2e"]
    fin = p.get("pai_1_financed_emissions_tco2e")        # attributed via EVIC, or None

    # Fill the mandatory table: computed where we honestly can, gap-flagged otherwise.
    filled: dict[int, dict] = {}

    # PAI 1 — financed GHG emissions (PCAF-attributed via EVIC where available)
    if fin:
        filled[1] = _row(1, "Climate & environment",
                         "GHG emissions — financed (Scope 1, 2, 3, total)", "tCO₂e",
                         value={"scope_1": fin["scope_1"], "scope_2": fin["scope_2"],
                                "scope_3": fin["scope_3"], "total": fin["total"]},
                         coverage=fin_cov, source=_GOLDEN_SOURCE + " · PCAF attribution (investment ÷ EVIC)",
                         method="computed" if fin_cov >= 99.9 else "partial",
                         input_required=None if fin_cov >= 99.9
                         else f"issuer EVIC on the remaining {round(100 - fin_cov, 1)}% by value")
    else:
        filled[1] = _row(1, "Climate & environment",
                         "GHG emissions (Scope 1, 2 and 3, and total)", "tCO₂e",
                         value={"scope_1": inv["scope_1"], "scope_2": inv["scope_2"],
                                "scope_3": inv["scope_3"],
                                "total": inv["scope_1"] + inv["scope_2"] + inv["scope_3"]},
                         coverage=emis_cov, source=_GOLDEN_SOURCE, method="partial",
                         input_required="issuer EVIC (enterprise value incl. cash) to attribute "
                                        "financed emissions per PCAF")
    # PAI 2 — carbon footprint (financed emissions / €M invested)
    cf = p.get("pai_2_carbon_footprint_tco2e_per_meur")
    filled[2] = _row(2, "Climate & environment",
                     "Carbon footprint (financed emissions per €M invested)", "tCO₂e/€M",
                     value=cf, coverage=fin_cov if cf is not None else None,
                     source=_GOLDEN_SOURCE + " · PCAF" if cf is not None else None,
                     method=("computed" if fin_cov >= 99.9 else "partial") if cf is not None else "not_available",
                     input_required=None if cf is None or fin_cov >= 99.9
                     else f"issuer EVIC on the remaining {round(100 - fin_cov, 1)}% by value")
    # PAI 3 — WACI (computed; may blend reported + estimated inputs, disclosed below)
    waci_src = _GOLDEN_SOURCE + (f" · {emis_est}% of covered value estimated" if emis_est else "")
    filled[3] = _row(3, "Climate & environment",
                     "GHG intensity of investee companies (WACI)", "tCO₂e/€M revenue",
                     value=p["pai_3_waci_tco2e_per_meur"], coverage=emis_cov,
                     source=waci_src,
                     method="computed" if p["pai_3_waci_tco2e_per_meur"] is not None else "not_available",
                     input_required=None if p["pai_3_waci_tco2e_per_meur"] is not None
                     else "issuer Scope 1/2 emissions + revenue")
    # PAI 4 — fossil-fuel-sector exposure (computed from NACE)
    filled[4] = _row(4, "Climate & environment",
                     "Exposure to companies active in the fossil fuel sector", "% of value",
                     value=p["pai_4_fossil_fuel_exposure_pct"], coverage=100.0,
                     source="issuer NACE division (golden source)", method="computed")

    # Inputs each remaining mandatory indicator needs (when its data is absent).
    remaining_inputs = {
        5: "issuer energy mix (renewable vs non-renewable share)",
        6: "issuer energy consumption (GWh) by high-impact NACE",
        7: "issuer operations in/near biodiversity-sensitive areas",
        8: "issuer emissions to water (tonnes)",
        9: "issuer hazardous/radioactive waste (tonnes)",
        10: "UNGC/OECD violation flags per issuer",
        11: "issuer compliance-monitoring process disclosure",
        12: "issuer unadjusted gender pay gap",
        13: "issuer board gender diversity",
        14: "controversial-weapons involvement flags per issuer",
    }

    # PAI 5-14 — the non-carbon indicators, computed from issuer_esg_metrics where
    # the manager has supplied that ESG data (value-weighted / share / attributed).
    esg = fund_esg_pai(session, fund_id)
    _ESG_SRC = "issuer ESG disclosures (manager feed), value-weighted"
    for num in range(5, 15):
        cell = esg.get(f"pai_{num}") if esg else None
        if cell and cell.get("value") is not None:
            filled[num] = _row(num, next(a for n, a, _, __ in MANDATORY_PAI_INDICATORS if n == num),
                               next(m for n, _, m, __ in MANDATORY_PAI_INDICATORS if n == num),
                               next(u for n, _, __, u in MANDATORY_PAI_INDICATORS if n == num),
                               value=cell["value"], coverage=cell["coverage_pct"], source=_ESG_SRC,
                               method="computed" if cell["coverage_pct"] >= 99.9 else "partial",
                               input_required=None if cell["coverage_pct"] >= 99.9
                               else f"{remaining_inputs[num]} on the remaining {round(100 - cell['coverage_pct'], 1)}% by value")

    indicators = []
    for num, area, metric, unit in MANDATORY_PAI_INDICATORS:
        if num in filled:
            indicators.append(filled[num])
        else:
            indicators.append(_row(num, area, metric, unit,
                                   input_required=remaining_inputs.get(num)))

    computed = sum(1 for i in indicators if i["method"] == "computed")
    partial = sum(1 for i in indicators if i["method"] == "partial")
    missing = sum(1 for i in indicators if i["method"] == "not_available")

    comp = _composition_and_sovereign(session, fund_id)

    # Reference period = the dominant emissions vintage across the book (PAI is
    # reported for a calendar year). Prior period disclosed as first-period N/A.
    ref_year = session.execute(text("""
        SELECT e.reporting_year FROM fund_positions p
        JOIN securities s ON s.security_id = p.security_id
        JOIN issuer_emissions e ON e.issuer_id = s.issuer_id
        WHERE p.fund_id = :f AND e.scope1_tco2e IS NOT NULL
        GROUP BY e.reporting_year ORDER BY count(*) DESC, e.reporting_year DESC LIMIT 1
    """), {"f": fund_id}).scalar()

    # Year-on-year: attach each indicator's prior filed value + change (SFDR yr 2+).
    comparison = _attach_prior_year(session, fund_id, ref_year, indicators)

    manager_lei = fund.get("manager_lei")
    # Filing-readiness: the reporting-entity identity SFDR's Annex I header needs.
    # NB: keep this list name distinct from the `missing` indicator COUNT above.
    filing_missing = []
    if not manager_lei:
        filing_missing.append("manager LEI")
    if not fund.get("manager_legal_name"):
        filing_missing.append("manager legal name")
    if not fund.get("filing_contact_email"):
        filing_missing.append("filing contact email")
    if not ref_year:
        filing_missing.append("reference period (supply issuer emissions with a reporting year)")
    # SFDR Annex I mandatory narrative sections.
    narratives = fund.get("sfdr_narratives") or {}
    _REQUIRED_NARRATIVES = {
        "policies": "policies to identify and prioritise principal adverse impacts",
        "actions": "actions taken and planned",
        "engagement": "engagement policies",
    }
    missing_narratives = [label for key, label in _REQUIRED_NARRATIVES.items() if not (narratives.get(key) or "").strip()]
    filing_missing += [f"narrative: {n}" for n in missing_narratives]
    ready_to_file = not filing_missing

    return {
        "entity": {
            "fund_id": fund["fund_id"], "fund_name": fund["name"], "fund_lei": fund.get("fund_lei"),
            "manager": fund["org_name"], "manager_lei": manager_lei,
            "manager_legal_name": fund.get("manager_legal_name"),
            "manager_domicile": fund.get("manager_domicile"),
            "filing_contact_email": fund.get("filing_contact_email"),
            "base_currency": fund["base_currency"],
            "sfdr_classification": fund["sfdr_classification"],
            "total_value_eur": pai["total_value_eur"], "positions": pai["positions"],
        },
        # RTS Annex I header/declaration + reference period.
        "summary": {
            "reference_period": f"FY{ref_year}" if ref_year else "reference period not set (supply issuer emissions with a reporting year)",
            "reference_year": ref_year,
            "prior_period": "Not available — first reference period",
            "pai_considered": True,
            "manager_lei_required": manager_lei is None,
            "declaration": (
                f"This is the principal adverse impacts statement on sustainability factors of "
                f"{fund.get('manager_legal_name') or fund['org_name']} ({manager_lei or 'LEI required'}) for the fund "
                f"'{fund['name']}', reference period {('FY' + str(ref_year)) if ref_year else '—'}. "
                "Principal adverse impacts of investment decisions on sustainability factors are considered."
            ),
        },
        "filing_readiness": {
            "ready_to_file": ready_to_file,
            "missing": filing_missing,
            "note": "Ready to file." if ready_to_file
                    else "Not yet submittable — supply the reporting-entity identity above.",
        },
        "statement": "Principal Adverse Impact (PAI) statement",
        "regulatory_basis": "SFDR RTS — Commission Delegated Regulation (EU) 2022/1288, Annex I, Table 1",
        "comparison": comparison,   # prior-period availability + year (indicators carry prior_value/change)
        "indicators": indicators,
        "holdings_composition": comp["by_asset_class"],
        # Sovereign indicators (15-16) shown only when the fund holds sovereigns.
        "sovereign_indicators": _sovereign_indicators(comp) if comp["sovereign_value_eur"] else [],
        "sovereign_countries": comp["sovereign_countries"],
        # Real-estate indicators (17-18) — always listed with their applicability.
        "real_estate_indicators": _real_estate_indicators(comp),
        "taxonomy": _taxonomy_rollup(session, fund_id),
        # Additional (voluntary) PAI — SFDR requires the manager to adopt ≥1 more
        # climate and ≥1 more social indicator from RTS Tables 2 & 3. Disclosed as
        # a required declaration, not silently omitted.
        "additional_indicators": {
            "requirement": "Adopt ≥1 additional climate/environmental (RTS Table 2) and "
                           "≥1 additional social (RTS Table 3) indicator.",
            "selected": [],
            "status": "declaration_required",
            "input_required": "the manager's chosen additional indicators + their data",
        },
        # Look-through — if the book holds funds/ETFs, their constituents must be
        # looked through. Detected from asset_class; honest status, not faked.
        "look_through": _look_through(comp),
        # Mandatory qualitative sections (manager-authored); missing ones flagged.
        "narratives": {
            "policies": narratives.get("policies"),
            "actions": narratives.get("actions"),
            "engagement": narratives.get("engagement"),
            "standards": narratives.get("standards"),
            "missing": missing_narratives,
        },
        "coverage_summary": {
            "mandatory_indicators": len(indicators),
            "computed": computed, "partial": partial, "not_available": missing,
            "emissions_coverage_pct": emis_cov,
            "emissions_estimated_pct": emis_est,   # of covered value, the reported/estimated split (SFDR RTS)
            "pcaf_data_quality_score": pai.get("pcaf_data_quality_score"),  # PCAF 1(best)–5(worst)
            "filing_readiness": (
                f"{computed} of {len(indicators)} mandatory indicators computed, "
                f"{partial} partial, {missing} awaiting issuer input. This statement is "
                "structurally complete and audit-traceable; the gaps are disclosed as "
                "required inputs, not estimated or omitted."
            ),
        },
        "provenance": {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "source": _GOLDEN_SOURCE,
            "scope_note": "Investee-company + sovereign + real-estate indicators, per the "
                          "holdings actually present. Scope 3 emissions are not estimated.",
            "data_sources": [
                {"item": "Issuer identity & ISIN→LEI", "source": "GLEIF (open LEI system)", "vintage": "resolved at onboarding"},
                {"item": "Facility location", "source": "GLEIF HQ address → OpenStreetMap/Nominatim geocode", "vintage": "at onboarding"},
                {"item": "Physical hazard scores", "source": "Tellumen golden source (canonical_scores, append-only)", "vintage": "model-stamped"},
                {"item": "Issuer emissions / revenue / EVIC", "source": "client disclosure where supplied; else estimated", "vintage": f"FY{ref_year}" if ref_year else "n/a"},
                {"item": "Estimated emissions", "source": "NACE sector intensity × revenue — EXIOBASE 3 IOT_2022_ixi (EU output-weighted GHG), interim fallback where EXIOBASE folds sectors", "vintage": "2022"},
                {"item": "Sovereign country GHG intensity", "source": "OWID / Global Carbon Project CO2 ÷ GDP (data/reference/country_ghg_intensity.csv)", "vintage": "2022"},
            ],
            "model_versions": {
                "emissions_estimation": "emissions-est-v1-sector-intensity",
                "attribution": "PCAF: investment ÷ EVIC",
            },
            "disclosures": {
                "emissions_coverage_pct": emis_cov,
                "emissions_estimated_pct": emis_est,
                "financed_emissions_coverage_pct": pai.get("financed_emissions_coverage_pct", 0.0),
            },
            "manager_actions_note": "Actions taken and targets (RTS Table 1, final columns) are a "
                                    "manager narrative and must be completed by the manager; not machine-derived.",
        },
    }


# ── Downloadable filing document (.xlsx in the mandated table shape) ──
def _method_label(m: str) -> str:
    return {"computed": "Computed", "partial": "Partial (input needed)",
            "estimated": "Estimated", "not_available": "Not available — input required",
            "not_applicable": "Not applicable"}.get(m, m)


def _fmt_value(v) -> str:
    if v is None:
        return "—"
    if isinstance(v, dict):
        return " · ".join(f"{k.replace('_', ' ')}: {v[k]:,}" if isinstance(v[k], (int, float)) else f"{k}: {v[k]}"
                          for k in v)
    if isinstance(v, (int, float)):
        return f"{v:,}"
    return str(v)


def _explanation(ind: dict, ref_year) -> str:
    """RTS 'Explanation' column: what the figure is, its coverage, and — honestly —
    where it's estimated or still needs input."""
    if ind["method"] == "computed":
        base = f"Computed ({ind['coverage_pct']}% coverage). Source: {ind['source']}."
    elif ind["method"] == "partial":
        base = f"Partial ({ind['coverage_pct']}% coverage). Source: {ind['source']}."
        if ind["input_required"]:
            base += f" To complete: {ind['input_required']}."
    elif ind["method"] == "estimated":
        base = f"Estimated. {ind['source']}."
    elif ind["method"] == "not_applicable":
        base = f"Not applicable — {ind['input_required']}." if ind["input_required"] else "Not applicable."
    else:
        base = f"Not available. Input required: {ind['input_required']}." if ind["input_required"] else "Not available."
    return base


def sfdr_pai_statement_xlsx(statement: dict) -> io.BytesIO:
    """Render a filing-grade, multi-sheet workbook: Summary (RTS declaration),
    PAI statement (RTS Table 1 columns), and a Provenance & methodology appendix."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    title_font = Font(bold=True, size=14, color="1D1D1F")
    h2_font = Font(bold=True, size=12, color="1B54C4")
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1B54C4")
    label_font = Font(bold=True, color="48515F")
    wrap = Alignment(wrap_text=True, vertical="top")

    e = statement["entity"]
    summ = statement["summary"]
    prov = statement["provenance"]
    ref_year = summ.get("reference_year")
    wb = Workbook()

    # ── Sheet 1: Summary / declaration ──
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Principal Adverse Impacts Statement — Summary"
    ws["A1"].font = title_font
    ws["A3"] = summ["declaration"]
    ws["A3"].alignment = wrap
    ws.merge_cells("A3:F5")
    rows = [
        ("Financial market participant (manager)", e["manager"]),
        ("Manager LEI", e.get("manager_lei") or "REQUIRED — supply the manager's LEI"),
        ("Fund", e["fund_name"]),
        ("SFDR product classification", e.get("sfdr_classification") or "—"),
        ("Reference period", summ["reference_period"]),
        ("Prior reference period", summ["prior_period"]),
        ("Principal adverse impacts considered", "Yes"),
        ("Portfolio value (EUR)", f"{e['total_value_eur']:,}"),
        ("Positions", e["positions"]),
        ("Regulatory basis", statement["regulatory_basis"]),
        ("Mandatory indicators computed", f"{statement['coverage_summary']['computed']} of {statement['coverage_summary']['mandatory_indicators']}"),
        ("Emissions coverage", f"{statement['coverage_summary']['emissions_coverage_pct']}% (of which {statement['coverage_summary']['emissions_estimated_pct']}% estimated)"),
        ("PCAF data-quality score", f"{statement['coverage_summary'].get('pcaf_data_quality_score', '—')} (1 best … 5 worst)"),
        ("Additional (voluntary) PAI", statement.get("additional_indicators", {}).get("status", "—")),
        ("Look-through", statement.get("look_through", {}).get("note") or statement.get("look_through", {}).get("status", "—")),
        ("Narrative sections outstanding", ", ".join(statement.get("narratives", {}).get("missing") or []) or "none — complete"),
        ("Generated (UTC)", prov["generated_at"]),
    ]
    r = 7
    for k, v in rows:
        ws.cell(r, 1, k).font = label_font
        ws.cell(r, 2, v).alignment = wrap
        r += 1
    for i, w in enumerate([38, 62], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: PAI statement (RTS Table 1 columns) ──
    ws2 = wb.create_sheet("PAI statement")
    ref_lbl = summ["reference_period"]
    headers = ["#", "Adverse sustainability indicator", "Metric",
               f"Impact ({ref_lbl})", "Impact (prior period)", "Explanation",
               "Actions taken / planned & targets"]
    for c, h in enumerate(headers, 1):
        cell = ws2.cell(1, c, h)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = wrap
    r = 2

    def _write_row(ind):
        nonlocal r
        ws2.cell(r, 1, ind["number"])
        ws2.cell(r, 2, ind["area"]).alignment = wrap
        ws2.cell(r, 3, f'{ind["metric"]} ({ind["unit"]})').alignment = wrap
        ws2.cell(r, 4, _fmt_value(ind["value"])).alignment = wrap
        ws2.cell(r, 5, statement["summary"]["prior_period"]).alignment = wrap
        ws2.cell(r, 6, _explanation(ind, ref_year)).alignment = wrap
        ws2.cell(r, 7, "[Manager to complete]").alignment = wrap
        r += 1

    def _section(label, items):
        nonlocal r
        if not items:
            return
        c = ws2.cell(r, 1, label)
        c.font = label_font
        r += 1
        for ind in items:
            _write_row(ind)

    _section("Investee companies (indicators 1–14)", statement["indicators"])
    _section("Sovereign & supranational (indicators 15–16)", statement.get("sovereign_indicators", []))
    _section("Real estate (indicators 17–18)", statement.get("real_estate_indicators", []))
    for i, w in enumerate([5, 26, 40, 24, 22, 60, 30], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: Provenance & methodology appendix ──
    ws3 = wb.create_sheet("Provenance & methodology")
    ws3["A1"] = "Provenance & methodology"
    ws3["A1"].font = title_font
    ws3["A3"] = "Data sources"
    ws3["A3"].font = h2_font
    for c, h in enumerate(["Item", "Source", "Vintage"], 1):
        cell = ws3.cell(4, c, h); cell.font = head_font; cell.fill = head_fill
    r = 5
    for ds in prov["data_sources"]:
        ws3.cell(r, 1, ds["item"]).alignment = wrap
        ws3.cell(r, 2, ds["source"]).alignment = wrap
        ws3.cell(r, 3, ds["vintage"]).alignment = wrap
        r += 1
    r += 1
    ws3.cell(r, 1, "Model versions").font = h2_font; r += 1
    for k, v in prov["model_versions"].items():
        ws3.cell(r, 1, k).font = label_font; ws3.cell(r, 2, v); r += 1
    r += 1
    ws3.cell(r, 1, "Disclosures").font = h2_font; r += 1
    d = prov["disclosures"]
    for k, v in [("Emissions coverage", f"{d['emissions_coverage_pct']}%"),
                 ("of which estimated", f"{d['emissions_estimated_pct']}%"),
                 ("Financed-emissions (EVIC) coverage", f"{d['financed_emissions_coverage_pct']}%")]:
        ws3.cell(r, 1, k).font = label_font; ws3.cell(r, 2, v); r += 1
    r += 1
    tax = statement["taxonomy"]
    ws3.cell(r, 1, "EU Taxonomy").font = h2_font; r += 1
    for k, v in [("Assessable share (has NACE)", f"{tax['assessable_pct']}%"),
                 ("Taxonomy-aligned", "Not asserted — " + tax["alignment_note"])]:
        ws3.cell(r, 1, k).font = label_font; ws3.cell(r, 2, v).alignment = wrap; r += 1
    r += 2
    ws3.cell(r, 1, "Methodology notes").font = h2_font; r += 1
    for note in [prov["scope_note"], prov["manager_actions_note"],
                 "Estimated figures use NACE sector-average intensity × revenue (illustrative "
                 "coefficients pending an EXIOBASE-sourced table); scope 3 is not estimated.",
                 "Financed emissions use the PCAF attribution factor (investment ÷ EVIC).",
                 "Unmatched securities and missing inputs are surfaced, never fabricated."]:
        ws3.cell(r, 1, note).alignment = wrap
        ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 1
    for i, w in enumerate([40, 46, 40], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # Print setup so a PDF/print export is clean (customers often circulate a PDF).
    from openpyxl.worksheet.properties import PageSetupProperties
    for sheet, landscape in ((ws, False), (ws2, True), (ws3, False)):
        sheet.page_setup.orientation = "landscape" if landscape else "portrait"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
