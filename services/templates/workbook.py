"""
Real formatted Excel templates/exports, shared across all 3 verticals -- one
function each, reused by bank.py/supply.py/insurance.py rather than each
router hand-rolling its own openpyxl code. A "market-standard template" means
more than a bare CSV header row: required vs optional fields marked, an
example row, and a field guide -- the same information a bank/insurer/agri
buyer's own onboarding doc would carry.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="0071E3", end_color="0071E3", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
EXAMPLE_FILL = PatternFill(start_color="F5F5F7", end_color="F5F5F7", fill_type="solid")
EXAMPLE_FONT = Font(italic=True, color="86868B")
REQUIRED_FONT = Font(bold=True, color="C2410C")


def _autosize(ws, ncols):
    for i in range(1, ncols + 1):
        col = get_column_letter(i)
        width = max((len(str(c.value)) for c in ws[col] if c.value is not None), default=10)
        ws.column_dimensions[col].width = min(max(width + 2, 12), 45)


def build_template_workbook(fields: list[dict]) -> io.BytesIO:
    """fields: [{name, required (bool), description, example}, ...]. Returns a
    2-sheet workbook: Data (header + one example row, greyed out) and Field Guide
    (name/required/description/example) -- delete-before-uploading is stated
    explicitly, not left implicit."""
    wb = Workbook()

    data_ws = wb.active
    data_ws.title = "Data"
    for i, f in enumerate(fields, start=1):
        cell = data_ws.cell(row=1, column=i, value=f["name"] + (" *" if f["required"] else ""))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, f in enumerate(fields, start=1):
        cell = data_ws.cell(row=2, column=i, value=f["example"])
        cell.fill = EXAMPLE_FILL
        cell.font = EXAMPLE_FONT
    data_ws.cell(row=3, column=1, value="^ EXAMPLE ROW — delete before uploading. Fields marked * are required.")
    data_ws.cell(row=3, column=1).font = Font(italic=True, size=9, color="86868B")
    _autosize(data_ws, len(fields))

    guide_ws = wb.create_sheet("Field Guide")
    headers = ["Field", "Required", "Description", "Example"]
    for i, h in enumerate(headers, start=1):
        cell = guide_ws.cell(row=1, column=i, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for r, f in enumerate(fields, start=2):
        guide_ws.cell(row=r, column=1, value=f["name"])
        req_cell = guide_ws.cell(row=r, column=2, value="Required" if f["required"] else "Optional")
        if f["required"]:
            req_cell.font = REQUIRED_FONT
        guide_ws.cell(row=r, column=3, value=f["description"])
        guide_ws.cell(row=r, column=4, value=f["example"])
        guide_ws.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")
    _autosize(guide_ws, len(headers))
    guide_ws.column_dimensions["C"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_export_workbook(headers: list[str], rows: list[list], sheet_name: str = "Data") -> io.BytesIO:
    """headers + rows -> a single-sheet formatted workbook, the Excel sibling of
    the existing client-side CSV export (same data, real spreadsheet formatting)."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for r, row in enumerate(rows, start=2):
        for i, val in enumerate(row, start=1):
            ws.cell(row=r, column=i, value=val)
    ws.freeze_panes = "A2"
    _autosize(ws, len(headers))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


SECTION_TITLE_FILL = PatternFill(start_color="1E2A44", end_color="1E2A44", fill_type="solid")
SECTION_TITLE_FONT = Font(bold=True, color="FFFFFF")
SECTION_COL_FONT = Font(bold=True, color="1E2A44")


def build_disclosure_workbook(data_headers: list[str], data_rows: list[list], data_sheet_name: str,
                              summary_blocks: list[dict]) -> io.BytesIO:
    """Two-sheet disclosure export: the per-entity data sheet PLUS a 'Computed disclosure' sheet built from the
    official-form annex sections, so the computed analytics (catastrophe/SCR/reinsurance/stranding/concentration/…)
    are in the downloadable filing, not just the per-row data. summary_blocks: [{title, columns, rows(list[list])}]."""
    wb = Workbook()
    ws = wb.active
    ws.title = data_sheet_name
    for i, h in enumerate(data_headers, start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for r, row in enumerate(data_rows, start=2):
        for i, val in enumerate(row, start=1):
            ws.cell(row=r, column=i, value=val)
    ws.freeze_panes = "A2"
    _autosize(ws, len(data_headers))

    sw = wb.create_sheet("Computed disclosure")
    maxcols = 1
    r = 1
    for blk in summary_blocks:
        tc = sw.cell(row=r, column=1, value=blk["title"])
        tc.fill = SECTION_TITLE_FILL
        tc.font = SECTION_TITLE_FONT
        r += 1
        cols = blk.get("columns") or []
        for i, h in enumerate(cols, start=1):
            hc = sw.cell(row=r, column=i, value=h)
            hc.font = SECTION_COL_FONT
        maxcols = max(maxcols, len(cols))
        r += 1
        for row in blk.get("rows", []):
            for i, val in enumerate(row, start=1):
                sw.cell(row=r, column=i, value=val)
            maxcols = max(maxcols, len(row))
            r += 1
        r += 1   # spacer between sections
    _autosize(sw, maxcols)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
